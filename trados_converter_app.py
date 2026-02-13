import streamlit as st
import xml.etree.ElementTree as ET
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def parse_trados_xml(xml_content):
    """Parse Trados XML analysis file and extract data"""
    try:
        root = ET.fromstring(xml_content)
        
        # Get project name
        project_node = root.find('.//project')
        project_name = project_node.get('name') if project_node is not None else "Unknown Project"
        
        # Get language
        language_node = root.find('.//language')
        language_name = language_node.get('name') if language_node is not None else "Unknown Language"
        
        # Get all files
        files = root.findall('.//file')
        
        data = []
        
        for idx, file_node in enumerate(files, start=1):
            file_name = file_node.get('name', 'Unknown')
            file_guid = file_node.get('guid', '')
            
            analyse = file_node.find('analyse')
            if analyse is None:
                continue
            
            # Extract match band data
            new_elem = analyse.find("new")
            fuzzy_50_74_elem = analyse.find("fuzzy[@min='50']")
            fuzzy_75_84_elem = analyse.find("fuzzy[@min='75']")
            fuzzy_85_94_elem = analyse.find("fuzzy[@min='85']")
            fuzzy_95_99_elem = analyse.find("fuzzy[@min='95']")
            exact_elem = analyse.find("exact")
            perfect_elem = analyse.find("perfect")
            ice_elem = analyse.find("inContextExact")
            repeated_elem = analyse.find("repeated")
            cross_file_elem = analyse.find("crossFileRepeated")
            locked_elem = analyse.find("locked")
            total_elem = analyse.find("total")
            
            # Get internal fuzzy matches
            internal_50_74_elem = analyse.find("internalFuzzy[@min='50']")
            internal_75_84_elem = analyse.find("internalFuzzy[@min='75']")
            internal_85_94_elem = analyse.find("internalFuzzy[@min='85']")
            internal_95_99_elem = analyse.find("internalFuzzy[@min='95']")
            
            # Extract word counts
            new = int(new_elem.get('words', 0)) if new_elem is not None else 0
            fuzzy_50_74 = int(fuzzy_50_74_elem.get('words', 0)) if fuzzy_50_74_elem is not None else 0
            fuzzy_75_84 = int(fuzzy_75_84_elem.get('words', 0)) if fuzzy_75_84_elem is not None else 0
            fuzzy_85_94 = int(fuzzy_85_94_elem.get('words', 0)) if fuzzy_85_94_elem is not None else 0
            fuzzy_95_99 = int(fuzzy_95_99_elem.get('words', 0)) if fuzzy_95_99_elem is not None else 0
            
            internal_50_74 = int(internal_50_74_elem.get('words', 0)) if internal_50_74_elem is not None else 0
            internal_75_84 = int(internal_75_84_elem.get('words', 0)) if internal_75_84_elem is not None else 0
            internal_85_94 = int(internal_85_94_elem.get('words', 0)) if internal_85_94_elem is not None else 0
            internal_95_99 = int(internal_95_99_elem.get('words', 0)) if internal_95_99_elem is not None else 0
            
            exact = int(exact_elem.get('words', 0)) if exact_elem is not None else 0
            perfect = int(perfect_elem.get('words', 0)) if perfect_elem is not None else 0
            ice = int(ice_elem.get('words', 0)) if ice_elem is not None else 0
            repeated = int(repeated_elem.get('words', 0)) if repeated_elem is not None else 0
            cross_file = int(cross_file_elem.get('words', 0)) if cross_file_elem is not None else 0
            locked = int(locked_elem.get('words', 0)) if locked_elem is not None else 0
            total = int(total_elem.get('words', 0)) if total_elem is not None else 0
            
            # Calculate grouped bands (matching VBA logic)
            imf_new = new + fuzzy_50_74 + fuzzy_75_84 + internal_50_74 + internal_75_84
            imf_low_fuzzy = fuzzy_85_94 + internal_85_94
            imf_high_fuzzy = fuzzy_95_99 + internal_95_99
            imf_100 = exact + perfect + ice
            imf_reps = repeated + cross_file
            imf_100_reps = imf_100 + imf_reps
            
            # Calculate weighted words (matching VBA formula)
            weighted = round(
                imf_new + 
                round(imf_low_fuzzy * 0.6, 0) + 
                round(imf_high_fuzzy * 0.4, 0) + 
                round(imf_100_reps * 0.33, 0),
                0
            )
            
            data.append({
                'ID': idx,
                'File Name': file_name,
                'No Match': imf_new,
                '85-94': imf_low_fuzzy,
                '95-99': imf_high_fuzzy,
                '100%': imf_100,
                'Reps': imf_reps,
                '100% Reps': imf_100_reps,
                'Total (Gross)': total,
                'Weighted (Net)': int(weighted)
            })
        
        return {
            'project_name': project_name,
            'language': language_name,
            'data': data
        }
    
    except Exception as e:
        st.error(f"Error parsing XML: {str(e)}")
        return None

def format_excel_sheet(ws, project_name, language, import_time):
    """Apply formatting to Excel sheet matching VBA styling"""
    
    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    
    # Style header rows
    ws['A1'] = f"Project: {project_name}"
    ws['A1'].font = Font(bold=True)
    
    ws['A2'] = f"Imported on {import_time}"
    
    # Style the data table header (row 3)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, 11):  # A to J
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add borders to data table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get the last row with data
    last_row = ws.max_row
    
    # Apply borders to all data cells
    for row in range(3, last_row + 1):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # Left-align file names
    for row in range(4, last_row):
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Highlight 100%, Reps, and 100% Reps columns with gray background
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for row in range(4, last_row):
        ws.cell(row=row, column=6).fill = gray_fill  # 100%
        ws.cell(row=row, column=7).fill = gray_fill  # Reps
        ws.cell(row=row, column=8).fill = gray_fill  # 100% Reps
    
    # Bold the Totals row
    totals_font = Font(bold=True)
    for col in range(1, 11):
        ws.cell(row=last_row, column=col).font = totals_font
    
    # Medium borders around the numeric data area (C to J)
    medium_border_left = Border(left=Side(style='medium'))
    medium_border_right = Border(right=Side(style='medium'))
    medium_border_top = Border(top=Side(style='medium'))
    medium_border_bottom = Border(bottom=Side(style='medium'))
    
    for row in range(3, last_row + 1):
        ws.cell(row=row, column=3).border = Border(
            left=Side(style='medium'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.cell(row=row, column=10).border = Border(
            right=Side(style='medium'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    for col in range(3, 11):
        ws.cell(row=3, column=col).border = Border(
            top=Side(style='medium'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        ws.cell(row=last_row, column=col).border = Border(
            bottom=Side(style='medium'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )

def create_excel_workbook(parsed_data_list):
    """Create Excel workbook with multiple sheets"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    import_time = datetime.now().strftime("%m/%d/%Y %I:%M:%S %p")
    
    for parsed_data in parsed_data_list:
        project_name = parsed_data['project_name']
        language = parsed_data['language']
        data = parsed_data['data']
        
        # Create sheet name from language (max 31 chars for Excel)
        sheet_name = language[:31] if len(language) <= 31 else language[:28] + "..."
        
        # Ensure unique sheet names
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name[:28]}_{counter}"
            counter += 1
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Add header info
        ws['A1'] = f"Project: {project_name}"
        ws['A2'] = f"Imported on {import_time}"
        
        # Add column headers
        headers = ['ID', 'File Name', 'No Match', '85-94', '95-99', '100%', 'Reps', '100% Reps', 'Total (Gross)', 'Weighted (Net)']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header)
        
        # Add data rows
        for row_idx, row_data in enumerate(data, start=4):
            ws.cell(row=row_idx, column=1, value=row_data['ID'])
            ws.cell(row=row_idx, column=2, value=row_data['File Name'])
            ws.cell(row=row_idx, column=3, value=row_data['No Match'])
            ws.cell(row=row_idx, column=4, value=row_data['85-94'])
            ws.cell(row=row_idx, column=5, value=row_data['95-99'])
            ws.cell(row=row_idx, column=6, value=row_data['100%'])
            ws.cell(row=row_idx, column=7, value=row_data['Reps'])
            ws.cell(row=row_idx, column=8, value=row_data['100% Reps'])
            ws.cell(row=row_idx, column=9, value=row_data['Total (Gross)'])
            ws.cell(row=row_idx, column=10, value=row_data['Weighted (Net)'])
        
        # Add totals row
        last_data_row = len(data) + 3
        totals_row = last_data_row + 1
        
        ws.cell(row=totals_row, column=2, value='Totals')
        
        # Sum formulas for columns C through J
        for col in range(3, 11):
            col_letter = get_column_letter(col)
            ws.cell(row=totals_row, column=col).value = f"=SUM({col_letter}4:{col_letter}{last_data_row})"
        
        # Apply formatting
        format_excel_sheet(ws, project_name, language, import_time)
    
    return wb

def main():
    st.set_page_config(page_title="Trados Analysis Converter", page_icon="📊", layout="wide")
    
    st.title("🌍 Trados XML Analysis to Excel Converter")
    st.markdown("Upload one or more Trados analysis XML files. Each language will be placed in a separate Excel tab.")
    
    st.divider()
    
    # File uploader
    uploaded_files = st.file_uploader(
        "Choose Trados XML analysis files",
        type=['xml'],
        accept_multiple_files=True,
        help="Upload one or more XML analysis files from Trados Studio"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} file(s) uploaded")
        
        # Display uploaded files
        with st.expander("📁 Uploaded Files"):
            for file in uploaded_files:
                st.write(f"• {file.name}")
        
        if st.button("🚀 Convert to Excel", type="primary", use_container_width=True):
            with st.spinner("Processing XML files..."):
                parsed_data_list = []
                
                for uploaded_file in uploaded_files:
                    xml_content = uploaded_file.read()
                    parsed_data = parse_trados_xml(xml_content)
                    
                    if parsed_data:
                        parsed_data_list.append(parsed_data)
                        st.info(f"✓ Parsed: **{parsed_data['language']}** from {uploaded_file.name}")
                
                if parsed_data_list:
                    # Create Excel workbook
                    wb = create_excel_workbook(parsed_data_list)
                    
                    # Save to BytesIO
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.success("✅ Excel file created successfully!")
                    
                    # Display summary
                    st.markdown("### 📊 Summary")
                    summary_data = []
                    for parsed_data in parsed_data_list:
                        total_files = len(parsed_data['data'])
                        total_words = sum(row['Total (Gross)'] for row in parsed_data['data'])
                        weighted_words = sum(row['Weighted (Net)'] for row in parsed_data['data'])
                        
                        summary_data.append({
                            'Language': parsed_data['language'],
                            'Files': total_files,
                            'Total Words': f"{total_words:,}",
                            'Weighted Words': f"{weighted_words:,}"
                        })
                    
                    st.dataframe(summary_data, use_container_width=True, hide_index=True)
                    
                    # Download button
                    project_name = parsed_data_list[0]['project_name']
                    safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    filename = f"{safe_name}_Analysis.xlsx" if safe_name else "Trados_Analysis.xlsx"
                    
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=excel_buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.error("❌ No valid analysis files could be parsed.")
    else:
        st.info("👆 Upload XML files to get started")
        
        # Show example
        with st.expander("ℹ️ How to use"):
            st.markdown("""
            1. **Export analysis** from Trados Studio as XML files
            2. **Upload** one or more XML files (one per target language)
            3. **Click Convert** to process
            4. **Download** the formatted Excel file with all languages in separate tabs
            
            **Features:**
            - ✅ Grouped match bands (No Match, 85-94, 95-99, 100%, Reps)
            - ✅ Weighted word count calculations
            - ✅ Multiple languages in separate tabs
            - ✅ Professional formatting matching your VBA macro
            """)

if __name__ == "__main__":
    main()
